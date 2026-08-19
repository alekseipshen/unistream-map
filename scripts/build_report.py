#!/usr/bin/env python3
"""Excel-отчёт по курсам вокруг каждого отделения. Два режима.

Внутренний (по умолчанию): с оценками — кто нас обыгрывает, где открыто окно
перепродажи, выделения красным.

Для руководства (--clean): только факты — какие банки рядом, по каким адресам,
какие у них курсы и когда обновлены. Без сравнений, выводов и подсветки.

Наши курсы в обоих случаях живые, из API Юнистрима на момент сборки.

  python3 build_report.py               # внутренний
  python3 build_report.py --clean       # для руководства
"""
import json
import re
import sys
import time
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MSK = timezone(timedelta(hours=3))

BLUE = "1A3E8C"
YELLOW = "FFD400"
RED = "C62828"
GREY = "67707F"
HEAD = PatternFill("solid", fgColor=BLUE)
SUBHEAD = PatternFill("solid", fgColor="E8EDF7")
WARN = PatternFill("solid", fgColor="FDECEC")
THIN = Side(style="thin", color="D5DBE5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def arr(fname, const):
    src = (ROOT / "data" / fname).read_text()
    return json.loads(re.search(rf"const {const} = (\[.*?\]);", src, re.S).group(1))


def obj(fname, const):
    src = (ROOT / "data" / fname).read_text()
    m = re.search(rf"const {const} = (\{{.*?\}});", src, re.S)
    return json.loads(m.group(1)) if m else {}


def live_rates(branch_id):
    req = urllib.request.Request(f"https://unistream.ru/api/poses/exchange/{branch_id}",
                                 headers={"Accept-Language": "ru", "User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        d = json.load(r)
    return {x["currency"]: {"buy": x["buyRate"], "sell": x["sellRate"],
                            "at": x["lastUpdated"]} for x in d["exchangeRates"]}


def collect():
    branches = arr("branches.js", "BRANCHES")
    comps = arr("competitors.js", "COMPETITORS")
    mf = arr("offices_mainfin.js", "OFFICES_MF")
    poi = arr("poi.js", "POI")
    bank_rates = obj("bankrates.js", "BANK_RATES")
    poi_bank = obj("bankrates.js", "POI_BANK")

    by_branch = {b["num"]: [] for b in branches}
    norates = {b["num"]: 0 for b in branches}
    for c in comps:
        for n in c["near"]:
            by_branch[n["num"]].append({"bank": c["bank"], "address": c["address"] or c["name"],
                                        "d": n["d"], "rates": c["rates"], "kind": "офис",
                                        "at": (c.get("at") or "")[:16].replace("T", " ")})
    for o in mf:
        for n in o["near"]:
            by_branch[n["num"]].append({"bank": o["bank"], "address": o["address"],
                                        "d": n["d"], "rates": o["rates"], "kind": "офис",
                                        "at": o.get("at") or ""})
    seen_city = {b["num"]: set() for b in branches}
    for p in poi:
        alias = poi_bank.get(p["n"])
        r = bank_rates.get(alias) if alias else None
        for n in p["near"]:
            if not r:
                norates[n["num"]] += 1
                continue
            key = r["name"].lower()
            if key in seen_city[n["num"]]:
                continue
            seen_city[n["num"]].add(key)
            by_branch[n["num"]].append({
                "bank": r["name"], "address": p.get("addr") or "", "d": n["d"],
                "rates": r["rates"], "at": r.get("at") or "",
                "kind": "банк (везде одинаково)" if r.get("same") else "банк (по городу)"})
    for num in by_branch:
        by_branch[num].sort(key=lambda x: x["d"])
    return branches, by_branch, norates


def style_header(ws, row, titles, widths):
    for i, (t, w) in enumerate(zip(titles, widths), start=1):
        c = ws.cell(row=row, column=i, value=t)
        c.fill = HEAD
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


SOURCE_LABEL = {
    "офис": "курс отделения",
    "банк (везде одинаково)": "курс банка",
    "банк (по городу)": "курс банка по городу",
}


def build_clean(wb, branches, by_branch, norates, ours, now):
    """Отчёт для руководства: только факты, без сравнений и подсветки."""
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Курсы обмена валют рядом с отделениями ЮНИСТРИМ — Москва"
    ws["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws["A2"] = (f"Банки и обменные пункты в радиусе 1 км от каждого отделения. "
                f"Данные на {now:%d.%m.%Y %H:%M} МСК.")
    ws["A2"].font = Font(size=9, color=GREY)
    ws["A3"] = ("Источники: banki.ru, mainfin.ru (курсы сторонних банков), "
                "OpenStreetMap (расположение точек), API ЮНИСТРИМ (наши курсы).")
    ws["A3"].font = Font(size=9, color=GREY)

    style_header(ws, 5, ["Отделение", "Адрес", "Метро", "Банков\nрядом",
                         "Без публикуемых\nкурсов", "USD у нас\nпокупка / продажа",
                         "EUR у нас\nпокупка / продажа", "USD рядом\nпокупка, мин–макс",
                         "USD рядом\nпродажа, мин–макс"],
                 [12, 40, 24, 10, 15, 18, 18, 20, 20])
    row = 6
    for b in branches:
        lst = by_branch[b["num"]]
        our = ours[b["num"]]
        buys = [c["rates"]["USD"]["buy"] for c in lst if c["rates"].get("USD", {}).get("buy")]
        sells = [c["rates"]["USD"]["sell"] for c in lst if c["rates"].get("USD", {}).get("sell")]
        rng = lambda v: f"{min(v):.2f} – {max(v):.2f}" if v else "—"
        pair = lambda cur: (f"{our[cur]['buy']} / {our[cur]['sell']}" if our.get(cur) else "—")
        vals = [f"№ {b['num']}", b["address"], ", ".join(b["metro"]) or "—",
                len(lst), norates[b["num"]], pair("USD"), pair("EUR"), rng(buys), rng(sells)]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=i in (2, 3))
            c.font = Font(size=10, bold=(i == 1))
        row += 1
    ws.freeze_panes = "A6"

    for b in branches:
        sh = wb.create_sheet(f"№{b['num']}")
        our = ours[b["num"]]
        sh["A1"] = f"Отделение № {b['num']} — {b['address']}"
        sh["A1"].font = Font(bold=True, size=13, color=BLUE)
        sh["A2"] = f"Метро: {', '.join(b['metro']) or '—'}. Банков в радиусе 1 км: {len(by_branch[b['num']])}."
        sh["A2"].font = Font(size=9, color=GREY)

        sh["A4"] = "Курсы отделения"
        sh["A4"].font = Font(bold=True, size=11)
        style_header(sh, 5, ["Валюта", "Покупка", "Продажа", "Обновлено"], [12, 12, 12, 18])
        r = 6
        for cur, v in our.items():
            for i, val in enumerate([cur, v["buy"], v["sell"], (v.get("at") or "")[11:16]], start=1):
                cell = sh.cell(row=r, column=i, value=val)
                cell.border = BOX
                cell.font = Font(size=10, bold=(i == 1))
            r += 1

        r += 1
        sh.cell(row=r, column=1, value="Курсы в банках рядом").font = Font(bold=True, size=11)
        r += 1
        style_header(sh, r, ["Банк", "Адрес", "Расстояние,\nм", "Данные", "Валюта",
                             "Покупка", "Продажа", "Обновлено"],
                     [24, 42, 11, 20, 9, 11, 11, 16])
        r += 1
        for c in by_branch[b["num"]]:
            first = True
            for cur in sorted(c["rates"]):
                t = c["rates"][cur]
                vals = [c["bank"] if first else "", c["address"] if first else "",
                        c["d"] if first else "", SOURCE_LABEL.get(c["kind"], c["kind"]) if first else "",
                        cur, t.get("buy"), t.get("sell"), c.get("at", "") if first else ""]
                for i, v in enumerate(vals, start=1):
                    cell = sh.cell(row=r, column=i, value=v)
                    cell.border = BOX
                    cell.alignment = Alignment(vertical="top", wrap_text=(i == 2))
                    cell.font = Font(size=10, bold=(i == 1 and first))
                if first:
                    sh.cell(row=r, column=1).fill = SUBHEAD
                first = False
                r += 1
        sh.freeze_panes = "A8"


def main():
    clean = "--clean" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    default_name = ("Курсы_рядом_с_отделениями" if clean else "Конкуренты_ЮНИСТРИМ")
    out_path = Path(args[0]) if args else \
        ROOT / f"report/{default_name}_{datetime.now(MSK):%Y-%m-%d}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    branches, by_branch, norates = collect()
    ours = {}
    for b in branches:
        try:
            ours[b["num"]] = live_rates(b["id"])
        except Exception as e:
            print(f"  {b['num']}: живые курсы недоступны ({e}), берём сохранённые", file=sys.stderr)
            ours[b["num"]] = {r["cur"]: {"buy": r["buy"], "sell": r["sell"], "at": r.get("at")}
                              for r in b["rates"]}
        time.sleep(0.8)

    wb = Workbook()
    now = datetime.now(MSK)
    if clean:
        build_clean(wb, branches, by_branch, norates, ours, now)
        wb.save(out_path)
        print(f"готово: {out_path}")
        print(f"листов: {len(wb.sheetnames)}, отделений: {len(branches)}")
        return

    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Конкурентное окружение отделений ЮНИСТРИМ — Москва"
    ws["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws["A2"] = (f"Радиус 1 км. Сформировано {now:%d.%m.%Y %H:%M} МСК. "
                f"Источники курсов: banki.ru и mainfin.ru, наши курсы — API ЮНИСТРИМ.")
    ws["A2"].font = Font(size=9, color=GREY)

    titles = ["Отделение", "Адрес", "Конкурентов\nв 1 км", "Из них\nвыгоднее нас",
              "Точек без\nкурсов", "USD у нас\nпок/прод", "Лучшая покупка\nрядом",
              "Дешевле всех\nпродают", "Окно перепродажи"]
    style_header(ws, 4, titles, [12, 42, 12, 12, 11, 16, 22, 22, 34])
    row = 5
    for b in branches:
        lst = by_branch[b["num"]]
        our = ours[b["num"]]
        better = 0
        for c in lst:
            for cur, t in c["rates"].items():
                o = our.get(cur)
                if o and ((t.get("buy") and t["buy"] > o["buy"]) or
                          (t.get("sell") and t["sell"] < o["sell"])):
                    better += 1
                    break
        usd = our.get("USD")
        best_buy = max((c for c in lst if c["rates"].get("USD", {}).get("buy")),
                       key=lambda c: c["rates"]["USD"]["buy"], default=None)
        best_sell = min((c for c in lst if c["rates"].get("USD", {}).get("sell")),
                        key=lambda c: c["rates"]["USD"]["sell"], default=None)
        arb = ""
        if usd and best_sell and best_sell["rates"]["USD"]["sell"] < usd["buy"]:
            diff = usd["buy"] - best_sell["rates"]["USD"]["sell"]
            arb = (f"USD: покупаем по {usd['buy']}, {best_sell['bank']} "
                   f"продаёт по {best_sell['rates']['USD']['sell']} "
                   f"({best_sell['d']} м) — разница {diff:.2f} ₽")
        vals = [
            f"№ {b['num']}", b["address"], len(lst), better, norates[b["num"]],
            f"{usd['buy']} / {usd['sell']}" if usd else "—",
            f"{best_buy['rates']['USD']['buy']} — {best_buy['bank']}" if best_buy else "—",
            f"{best_sell['rates']['USD']['sell']} — {best_sell['bank']}" if best_sell else "—",
            arb or "нет",
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=i in (2, 7, 8, 9))
            c.font = Font(size=10, bold=(i == 1))
            if arb and i == 9:
                c.fill = WARN
                c.font = Font(size=10, color=RED, bold=True)
        row += 1
    ws.freeze_panes = "A5"

    for b in branches:
        sh = wb.create_sheet(f"№{b['num']}")
        our = ours[b["num"]]
        sh["A1"] = f"Отделение № {b['num']} — {b['address']}"
        sh["A1"].font = Font(bold=True, size=13, color=BLUE)
        metro = ", ".join(b["metro"]) if b["metro"] else "—"
        sh["A2"] = f"Метро: {metro}. Конкурентов в 1 км: {len(by_branch[b['num']])}."
        sh["A2"].font = Font(size=9, color=GREY)

        sh["A4"] = "Наши курсы"
        sh["A4"].font = Font(bold=True, size=11)
        style_header(sh, 5, ["Валюта", "Покупка", "Продажа", "Обновлено"], [12, 12, 12, 18])
        r = 6
        for cur, v in our.items():
            at = (v.get("at") or "")[11:16]
            for i, val in enumerate([cur, v["buy"], v["sell"], at], start=1):
                cell = sh.cell(row=r, column=i, value=val)
                cell.border = BOX
                cell.font = Font(size=10, bold=(i == 1))
            r += 1

        r += 1
        sh.cell(row=r, column=1, value="Конкуренты рядом").font = Font(bold=True, size=11)
        r += 1
        style_header(sh, r, ["Банк", "Адрес", "Метров", "Точность курса", "Валюта",
                             "Покупка", "Продажа", "У нас пок/прод", "Выгоднее нас"],
                     [22, 40, 9, 22, 9, 11, 11, 16, 26])
        r += 1
        for c in by_branch[b["num"]]:
            first = True
            for cur in sorted(c["rates"]):
                t = c["rates"][cur]
                o = our.get(cur)
                notes = []
                if o and t.get("buy") and t["buy"] > o["buy"]:
                    notes.append("покупает дороже")
                if o and t.get("sell") and t["sell"] < o["sell"]:
                    notes.append("продаёт дешевле")
                vals = [c["bank"] if first else "", c["address"] if first else "",
                        c["d"] if first else "", c["kind"] if first else "",
                        cur, t.get("buy"), t.get("sell"),
                        f"{o['buy']} / {o['sell']}" if o else "нет",
                        ", ".join(notes)]
                for i, v in enumerate(vals, start=1):
                    cell = sh.cell(row=r, column=i, value=v)
                    cell.border = BOX
                    cell.alignment = Alignment(vertical="top", wrap_text=(i == 2))
                    cell.font = Font(size=10, bold=(i == 1 and first))
                    if notes and i in (6, 7, 9):
                        cell.font = Font(size=10, color=RED, bold=True)
                if first:
                    sh.cell(row=r, column=1).fill = SUBHEAD
                first = False
                r += 1
        sh.freeze_panes = "A8"

    wb.save(out_path)
    print(f"готово: {out_path}")
    print(f"листов: {len(wb.sheetnames)}, отделений: {len(branches)}")


if __name__ == "__main__":
    main()
